"""
Market stats parser for the SF Market Stats Drive folder.

Pulls three files by name pattern:
  - Active and Pending Data <MMDDYYYY>.xlsx -> current active and pending counts
  - <range> RASE DATA.xlsx -> recently sold data with median price, DOM, SP%LP
  - 2026 MARKET STAT_as of <month> <year>.pdf -> monthly history for the chart

The xlsx files share the same MLS export schema. Sheet 1 row 1 is the header.
Status column drives filtering (Active, Pending, Contingent, Sold, etc).
SP%LP column gives the list-to-sale ratio per row.
"""

from __future__ import annotations

import io
import logging
import re
import statistics
from datetime import datetime, timedelta, timezone

logger = logging.getLogger(__name__)


def _norm(s):
    return (s or "").strip().lower()


def _to_float(value):
    """Parse a money or percent or numeric cell. Returns float or None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(value):
    """Parse a date cell from xlsx. Handles datetime, date, Excel float serial,
    and many common string formats. Returns datetime or None."""
    import datetime as _dt
    if value is None:
        return None
    # Native datetime
    if isinstance(value, _dt.datetime):
        return value
    # Native date (no time)
    if isinstance(value, _dt.date):
        return _dt.datetime(value.year, value.month, value.day)
    # Excel serial (days since 1899-12-30)
    if isinstance(value, (int, float)) and value > 1000:
        try:
            return _dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(value))
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    formats = (
        "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S",
        "%m-%d-%Y", "%m/%d/%Y %H:%M:%S", "%m/%d/%y %H:%M:%S",
        "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y", "%b %d, %Y", "%B %d, %Y",
    )
    for fmt in formats:
        try:
            return _dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_xlsx_rows(content_bytes):
    """Parse xlsx bytes into a list of dicts keyed by header row 1 of Sheet 1."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = []
    out = []
    for i, row in enumerate(rows_iter):
        if i == 0:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue
        if not any(row):
            continue
        d = {}
        for j, val in enumerate(row):
            if j < len(headers):
                d[headers[j]] = val
        out.append(d)
    return out


def find_file_by_pattern(files, pattern_substrings, mime_filter=None,
                          alt_substrings=None):
    """Return the most recently modified file whose name contains all
    `pattern_substrings`, optionally matching a mime_filter ('sheet' or 'pdf').

    `alt_substrings` is a fallback list of alternates: if no file matches
    `pattern_substrings`, try this second pattern set."""
    def matches(f, subs):
        name = (f.get("name") or "").lower()
        mime = (f.get("mimeType") or "").lower()
        if mime_filter and mime_filter not in mime:
            return False
        return all(p.lower() in name for p in subs)

    candidates = [f for f in files if matches(f, pattern_substrings)]
    if not candidates and alt_substrings:
        candidates = [f for f in files if matches(f, alt_substrings)]
    if not candidates:
        return None

    def sort_key(f):
        return f.get("modifiedTime") or f.get("name") or ""

    candidates.sort(key=sort_key, reverse=True)
    return candidates[0]


def summarize_active_pending(rows):
    """Count actives and pendings (incl. contingent) from the export."""
    active = 0
    pending = 0
    for r in rows:
        status = _norm(r.get("Status"))
        if status == "active":
            active += 1
        elif status in ("pending", "contingent", "active under contract"):
            pending += 1
    return {"active": active, "pending": pending}


def summarize_rase(rows, week_window_days=7):
    """Compute solds-this-week stats. RASE DATA is a year-to-date sold export
    so we filter by Selling Date within the last week_window_days. If fewer
    than 3 rows fall in the window (likely date parse failure or sparse week),
    we fall back to the most recent 30 rows by date."""
    SOLD_STATUSES = {"sold", "sld", "closed", "clsd", "cld"}
    parsed = []
    parse_failures = 0
    raw_samples = []
    skipped_status = 0
    for r in rows:
        status = _norm(r.get("Status"))
        if status and status not in SOLD_STATUSES:
            skipped_status += 1
            continue
        raw = r.get("Selling Date")
        sd = _to_date(raw)
        if sd is None and raw not in (None, "", 0):
            parse_failures += 1
            if len(raw_samples) < 3:
                raw_samples.append((repr(raw), type(raw).__name__))
        parsed.append((r, sd))
    logger.info(
        "RASE status filter: kept %d sold rows, skipped %d non-sold rows",
        len(parsed), skipped_status,
    )
    if parse_failures:
        logger.warning(
            "RASE date parse: %d failures out of %d rows. Sample raw values: %s",
            parse_failures, len(rows), raw_samples,
        )

    dated = [(r, d) for r, d in parsed if d is not None]
    dated.sort(key=lambda t: t[1], reverse=True)

    cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=week_window_days)
    week_rows = [r for r, d in dated if d >= cutoff]

    if len(week_rows) < 3 and dated:
        # Sparse week or date issues. Use most recent 30 sold rows.
        logger.info(
            "RASE week filter caught only %d rows. Falling back to most recent 30.",
            len(week_rows),
        )
        week_rows = [r for r, _ in dated[:30]]

    sold_count = len(week_rows)
    logger.info(
        "RASE summarize: %d rows total, %d parseable dates, %d in week (%dd window), final %d",
        len(rows), len(dated), len([r for r, d in dated if d >= cutoff]),
        week_window_days, sold_count,
    )

    sale_prices = [_to_float(r.get("Selling Price")) for r in week_rows]
    sale_prices = [v for v in sale_prices if v]
    median_price = int(statistics.median(sale_prices)) if sale_prices else None

    doms = [_to_float(r.get("Days on Market as Active")) for r in week_rows]
    doms = [v for v in doms if v is not None]
    median_dom = int(round(statistics.median(doms))) if doms else None

    sp_lp = [_to_float(r.get("SP%LP")) for r in week_rows]
    sp_lp = [v for v in sp_lp if v is not None]
    # Some exports store percent as 0.978 (0.978 = 97.8%), others as 97.8.
    # Normalize: any value <= 1.5 is treated as a fraction.
    if sp_lp:
        normalized = [v * 100 if v <= 1.5 else v for v in sp_lp]
        avg_sp_lp = round(statistics.mean(normalized), 1)
    else:
        avg_sp_lp = None

    return {
        "sold_count": sold_count,
        "median_price": median_price,
        "median_dom": median_dom,
        "avg_sp_lp_pct": avg_sp_lp,
    }


_MONTH_TOKENS = [
    ("january", "Jan"), ("february", "Feb"), ("march", "Mar"),
    ("april", "Apr"), ("may", "May"), ("june", "Jun"),
    ("july", "Jul"), ("august", "Aug"), ("september", "Sep"),
    ("october", "Oct"), ("november", "Nov"), ("december", "Dec"),
]


def extract_monthly_chart_from_pdf(content_bytes):
    """Extract a list of {month, value} dicts representing monthly median sale
    price for the most recent ~12 months. Falls back to [] on parse failure.

    The 2026 MARKET STAT pdf appends new monthly pages to the end. We scan all
    text, look for 'Median Sale Price' patterns near month/year tokens, and
    keep the latest 12.
    """
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(content_bytes))
        pages_text = []
        for page in reader.pages:
            try:
                t = page.extract_text() or ""
            except Exception:
                t = ""
            pages_text.append(t)
        full = "\n".join(pages_text)
    except Exception as exc:
        logger.warning("PDF read failed: %s", exc)
        return []

    results = []
    pattern = re.compile(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+(\d{4})"
        r"[\s\S]{0,200}?median\s+sale\s+price[:\s]*\$?([\d,\.]+)",
        re.IGNORECASE,
    )
    for m in pattern.finditer(full):
        month_short = m.group(1)[:3].title()
        year = m.group(2)
        price_str = m.group(3).replace(",", "")
        try:
            price = float(price_str)
            if price > 1000:
                price = round(price / 1000)
        except ValueError:
            continue
        label = f"{month_short} {year[-2:]}"
        results.append({"month": label, "value": int(price)})

    seen = set()
    deduped = []
    for r in reversed(results):
        key = r["month"]
        if key in seen:
            continue
        seen.add(key)
        deduped.insert(0, r)

    return deduped[-12:]


def build_market_stats_section(google_drive, folder_id, last_week_pending=None):
    """Top-level builder. Returns the marketStats dict for reports.json."""
    fallback = {
        "headline": "Sioux Falls market stats this week",
        "bullets": ["Source unavailable this week, check back next Sunday."],
        "chartTitle": "Median sale price, 12-month view",
        "chart": [],
        "sourceUrl": "",
    }

    try:
        files = google_drive.list_files_in_folder(folder_id)
    except Exception as exc:
        logger.warning("market stats folder list failed: %s", exc)
        return fallback

    if not files:
        return fallback

    ap_file = find_file_by_pattern(
        files,
        ["active and pending data"],
        mime_filter="sheet",
        alt_substrings=["active and pending"],
    )
    rase_file = find_file_by_pattern(files, ["rase data"], mime_filter="sheet")
    monthly_pdf = find_file_by_pattern(
        files,
        ["market stat", "2026"],
        mime_filter="pdf",
        alt_substrings=["market stat"],
    )

    logger.info(
        "market stats files: ap=%s rase=%s monthly_pdf=%s (folder had %d files)",
        ap_file.get("name") if ap_file else None,
        rase_file.get("name") if rase_file else None,
        monthly_pdf.get("name") if monthly_pdf else None,
        len(files),
    )

    actives = pendings = sold_count = 0
    median_price = median_dom = avg_sp_lp = None

    if ap_file:
        try:
            data = google_drive.download_file(ap_file["id"])
            if isinstance(data, str):
                data = data.encode()
            rows = parse_xlsx_rows(bytes(data))
            ap = summarize_active_pending(rows)
            actives = ap["active"]
            pendings = ap["pending"]
        except Exception as exc:
            logger.warning("active/pending parse failed: %s", exc)

    if rase_file:
        try:
            data = google_drive.download_file(rase_file["id"])
            if isinstance(data, str):
                data = data.encode()
            rows = parse_xlsx_rows(bytes(data))
            r = summarize_rase(rows)
            sold_count = r["sold_count"]
            median_price = r["median_price"]
            median_dom = r["median_dom"]
            avg_sp_lp = r["avg_sp_lp_pct"]
        except Exception as exc:
            logger.warning("RASE parse failed: %s", exc)

    chart = []
    if monthly_pdf:
        try:
            data = google_drive.download_file(monthly_pdf["id"])
            if isinstance(data, str):
                data = data.encode()
            chart = extract_monthly_chart_from_pdf(bytes(data))
        except Exception as exc:
            logger.warning("monthly PDF parse failed: %s", exc)

    # Bullet 1: inventory snapshot
    bullets = []
    bullets.append(f"Active: {actives}. Pending: {pendings}. Sold this week: {sold_count}.")

    # Bullet 2: sold price story
    if median_price is not None:
        parts = [f"Median sale price: ${median_price:,}"]
        if median_dom is not None:
            parts.append(f"average days on market: {median_dom}")
        if avg_sp_lp is not None:
            parts.append(f"list to sale ratio: {avg_sp_lp}%")
        bullets.append(". ".join(parts) + ".")
    else:
        bullets.append("No closed sales recorded for this week yet.")

    # Bullet 3: trend (vs last week's pending count if we have it)
    if last_week_pending is not None and pendings:
        delta = pendings - last_week_pending
        if delta > 0:
            bullets.append(f"Pending count is up by {delta} versus last week. Demand is leaning warmer.")
        elif delta < 0:
            bullets.append(f"Pending count is down by {abs(delta)} versus last week. Watch for buyer hesitation.")
        else:
            bullets.append("Pending count is flat versus last week. Steady week.")

    return {
        "headline": "Sioux Falls market stats this week",
        "bullets": bullets,
        "chartTitle": "Median sale price, 12-month view",
        "chart": chart,
        "sourceUrl": "",
        "_internal": {"pending_count": pendings},
    }
